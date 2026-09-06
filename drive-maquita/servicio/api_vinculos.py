# -*- coding: utf-8 -*-
"""Vínculos de datos del Almacén Maquita — "IMPORTRANGE" materializado.
====================================================================
Copia un rango de un Excel ORIGEN a una celda de un Excel DESTINO como VALORES
planos (no fórmula). Así el Destino muestra siempre el dato fresco SIN el mensaje
de "actualizar vínculos externos" de OnlyOffice (que no se puede desactivar).

Disparos de actualización:
  1) Automático: cuando se GUARDA el Origen (lo llama el callback de OnlyOffice,
     ver api_onlyoffice) -> refrescar_por_origen().
  2) Manual: botón "Actualizar datos vinculados" en el explorador
     -> POST /api/almacen/vinculos/actualizar?ruta=<destino>.

Lee el Origen con openpyxl data_only=True (valores calculados que OnlyOffice deja
al guardar) y reescribe el Destino con nucleo.subir() -> versionado y dedup.

Autoría: Equipo de Tecnología Maquita — 2026-07-24
"""
import io
import logging
import os

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from flask import Blueprint, jsonify, request

import almacen_bd as bd
import nucleo_archivos as nucleo
from api_archivos import error, usuario_actual
from seguridad_rutas import RutaInvalida, normalizar_ruta_virtual, ruta_fisica

log = logging.getLogger('almacen.vinculos')

bp_vinculos = Blueprint('almacen_vinculos', __name__)

EXT_EXCEL = {'xlsx', 'xlsm'}


def asegurar_esquema_vinculos():
    """Crea la tabla de vínculos si no existe. Idempotente."""
    with bd.conexion() as con:
        with con.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS vinculos_datos (
                    id SERIAL PRIMARY KEY,
                    origen_usuario   INTEGER NOT NULL,
                    origen_ruta      TEXT NOT NULL,
                    origen_hoja      TEXT NOT NULL,
                    origen_rango     TEXT NOT NULL,     -- ej "A1:C5"
                    destino_usuario  INTEGER NOT NULL,
                    destino_ruta     TEXT NOT NULL,
                    destino_hoja     TEXT NOT NULL,
                    destino_celda    TEXT NOT NULL,     -- esquina sup-izq, ej "A1"
                    creado_por       INTEGER,
                    creado_en        TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                    actualizado_en   TIMESTAMPTZ,
                    activo           BOOLEAN NOT NULL DEFAULT TRUE
                );
                CREATE INDEX IF NOT EXISTS ix_vinc_origen
                    ON vinculos_datos(origen_usuario, origen_ruta) WHERE activo;
                CREATE INDEX IF NOT EXISTS ix_vinc_destino
                    ON vinculos_datos(destino_usuario, destino_ruta) WHERE activo;
            """)


# ---------------------------------------------------------------------------
# Permisos (vínculos ENTRE usuarios)
# ---------------------------------------------------------------------------
def _correo_usuario(usuario):
    filas = bd.consultar('SELECT email FROM usuarios WHERE id = %s', (int(usuario),),
                         nomina=True)
    return (filas[0]['email'] or '').strip().lower() if filas else ''


def puede_leer(usuario, propietario, ruta):
    """¿`usuario` puede leer `ruta` de `propietario`? Es suyo, es master, o se lo
    compartieron a su correo y el compartido sigue vigente.

    Se comprueba TAMBIÉN en cada refresco: si le revocan o vence el compartido,
    el vínculo deja de traer datos (no se filtra información)."""
    usuario, propietario = int(usuario), int(propietario)
    if usuario == propietario:
        return True
    try:
        if bd.es_master(usuario):
            return True
    except Exception:
        pass
    correo = _correo_usuario(usuario)
    if not correo:
        return False
    filas = bd.consultar(
        'SELECT 1 FROM compartidos WHERE propietario_id = %s AND ruta = %s '
        'AND LOWER(email) = %s AND (expira_en IS NULL OR expira_en > NOW()) LIMIT 1',
        (propietario, ruta, correo))
    return bool(filas)


# ---------------------------------------------------------------------------
# Motor de refresco
# ---------------------------------------------------------------------------
def _leer_rango(usuario, ruta, hoja, rango):
    """Lee los valores calculados de un rango del Origen -> matriz de filas."""
    fis = ruta_fisica(usuario, ruta)
    wb = openpyxl.load_workbook(fis, data_only=True, read_only=True)
    try:
        if hoja not in wb.sheetnames:
            raise ValueError('La hoja "%s" no existe en el origen' % hoja)
        ws = wb[hoja]
        min_c, min_r, max_c, max_r = range_boundaries(rango)
        matriz = []
        for r in range(min_r, max_r + 1):
            fila = []
            for c in range(min_c, max_c + 1):
                fila.append(ws.cell(row=r, column=c).value)
            matriz.append(fila)
        return matriz
    finally:
        wb.close()


def _escribir_matriz(usuario, ruta, hoja, celda, matriz):
    """Escribe la matriz como VALORES en el Destino (a partir de `celda`) y lo
    vuelve a subir versionado con nucleo.subir()."""
    fis = ruta_fisica(usuario, ruta)
    wb = openpyxl.load_workbook(fis)  # con fórmulas/formato existentes
    if hoja not in wb.sheetnames:
        ws = wb.create_sheet(hoja)
    else:
        ws = wb[hoja]
    fila0, col0 = coordinate_to_tuple(celda)   # (row, col)
    for i, fila in enumerate(matriz):
        for j, valor in enumerate(fila):
            ws.cell(row=fila0 + i, column=col0 + j, value=valor)
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)
    carpeta, _, nombre = ruta.rpartition('/')
    nucleo.subir(usuario, carpeta or '/', nombre, buf)


def _refrescar(v):
    """Refresca UN vínculo (dict de la BD). Devuelve (ok, mensaje)."""
    try:
        # Revalidar permiso: si le revocaron/venció el compartido, no traer nada.
        if not puede_leer(v['destino_usuario'], v['origen_usuario'], v['origen_ruta']):
            log.warning('vinculo %s: sin permiso vigente sobre el origen', v.get('id'))
            return False, ('Ya no tienes permiso sobre el archivo de origen '
                           '(el compartido fue revocado o venció)')
        matriz = _leer_rango(v['origen_usuario'], v['origen_ruta'],
                             v['origen_hoja'], v['origen_rango'])
        _escribir_matriz(v['destino_usuario'], v['destino_ruta'],
                         v['destino_hoja'], v['destino_celda'], matriz)
        bd.ejecutar('UPDATE vinculos_datos SET actualizado_en = NOW() WHERE id = %s',
                    (v['id'],))
        return True, 'ok'
    except Exception as excepcion:
        log.error('vinculo %s: %s', v.get('id'), excepcion)
        return False, str(excepcion)


def refrescar_por_origen(usuario, ruta):
    """Refresca todos los vínculos cuyo ORIGEN es (usuario, ruta). Lo invoca el
    callback de OnlyOffice al guardar el origen. Nunca lanza (best-effort)."""
    try:
        ruta = normalizar_ruta_virtual(ruta)
        filas = bd.consultar(
            'SELECT * FROM vinculos_datos WHERE activo AND origen_usuario = %s '
            'AND origen_ruta = %s', (int(usuario), ruta))
        n = 0
        for v in filas:
            ok, _ = _refrescar(v)
            n += 1 if ok else 0
        if filas:
            log.info('refresco por origen %s: %s/%s vínculos', ruta, n, len(filas))
        return n
    except Exception as excepcion:
        log.warning('refrescar_por_origen %s: %s', ruta, excepcion)
        return 0


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
def _es_excel(ruta):
    ext = ruta.rsplit('.', 1)[-1].lower() if '.' in ruta else ''
    return ext in EXT_EXCEL


@bp_vinculos.route('/vinculos/crear', methods=['POST'])
def crear():
    """Crea un vínculo y lo materializa de una vez.
    JSON: origen_ruta, origen_hoja, origen_rango, destino_ruta, destino_hoja,
          destino_celda."""
    usuario = usuario_actual()
    d = request.get_json(silent=True) or {}
    try:
        o_ruta = normalizar_ruta_virtual(d.get('origen_ruta', ''))
        de_ruta = normalizar_ruta_virtual(d.get('destino_ruta', ''))
    except RutaInvalida as excepcion:
        return error(str(excepcion), excepcion.codigo)
    o_hoja = (d.get('origen_hoja') or '').strip()
    o_rango = (d.get('origen_rango') or '').strip().upper()
    de_hoja = (d.get('destino_hoja') or '').strip()
    de_celda = (d.get('destino_celda') or '').strip().upper()
    if not (o_ruta and de_ruta and o_hoja and o_rango and de_hoja and de_celda):
        return error('Faltan datos del vínculo', 400)
    if not (_es_excel(o_ruta) and _es_excel(de_ruta)):
        return error('Origen y destino deben ser hojas de cálculo (.xlsx)', 400)
    try:
        range_boundaries(o_rango)
        coordinate_to_tuple(de_celda)
    except Exception:
        return error('Rango o celda con formato inválido (ej "A1:C5" y "A1")', 400)

    # El ORIGEN puede ser de OTRA persona si se lo compartió (o si soy master).
    # El DESTINO siempre es un archivo propio: solo escribo en mi Drive.
    try:
        o_usuario = int(d.get('origen_usuario') or usuario)
    except (TypeError, ValueError):
        return error('Origen inválido', 400)
    if not puede_leer(usuario, o_usuario, o_ruta):
        return error('No tienes permiso sobre el archivo de origen. Pide que te lo '
                     'compartan desde el Drive.', 403)
    # El vinculo se materializa: ESCRIBE en el archivo de destino. Que sea
    # "mio" no basta si vive en una carpeta donde solo puedo leer.
    from permisos_accion import puede_escribir as _puede_escribir, carpeta_de, MOTIVO_LECTOR
    if not _puede_escribir(usuario, carpeta_de(de_ruta)):
        return error(MOTIVO_LECTOR, 403)

    fila = bd.ejecutar(
        """INSERT INTO vinculos_datos
           (origen_usuario, origen_ruta, origen_hoja, origen_rango,
            destino_usuario, destino_ruta, destino_hoja, destino_celda, creado_por)
           VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *""",
        (o_usuario, o_ruta, o_hoja, o_rango, usuario, de_ruta, de_hoja, de_celda,
         usuario))
    ok, msg = _refrescar(dict(fila))
    if not ok:
        return jsonify({'success': True, 'id': fila['id'],
                        'aviso': 'Vínculo creado, pero no se pudo traer los datos '
                                 'ahora: ' + msg})
    return jsonify({'success': True, 'id': fila['id'],
                    'mensaje': 'Vínculo creado y datos traídos'})


@bp_vinculos.route('/vinculos/listar', methods=['GET'])
def listar():
    """GET /vinculos/listar?ruta=<destino> — vínculos de ese archivo destino."""
    usuario = usuario_actual()
    try:
        ruta = normalizar_ruta_virtual(request.args.get('ruta', ''))
    except RutaInvalida as excepcion:
        return error(str(excepcion), excepcion.codigo)
    filas = bd.consultar(
        'SELECT id, origen_ruta, origen_hoja, origen_rango, destino_hoja, '
        'destino_celda, actualizado_en FROM vinculos_datos WHERE activo AND '
        'destino_usuario = %s AND destino_ruta = %s ORDER BY id', (usuario, ruta))
    return jsonify({'success': True, 'vinculos': filas})


@bp_vinculos.route('/vinculos/actualizar', methods=['POST'])
def actualizar():
    """POST /vinculos/actualizar?ruta=<destino> — refresca todos los vínculos de
    ese destino (botón manual)."""
    usuario = usuario_actual()
    try:
        ruta = normalizar_ruta_virtual(request.args.get('ruta', ''))
    except RutaInvalida as excepcion:
        return error(str(excepcion), excepcion.codigo)
    filas = bd.consultar(
        'SELECT * FROM vinculos_datos WHERE activo AND destino_usuario = %s AND '
        'destino_ruta = %s', (usuario, ruta))
    if not filas:
        return jsonify({'success': True, 'actualizados': 0,
                        'mensaje': 'Este archivo no tiene vínculos de datos'})
    n = sum(1 for v in filas if _refrescar(dict(v))[0])
    return jsonify({'success': True, 'actualizados': n, 'total': len(filas),
                    'mensaje': 'Datos actualizados (%s/%s)' % (n, len(filas))})


@bp_vinculos.route('/vinculos/origenes', methods=['GET'])
def origenes():
    """GET /vinculos/origenes — Excel que puedo usar como ORIGEN: los MÍOS y los
    que OTRAS personas me compartieron (vigentes). Cada uno trae su dueño."""
    usuario = usuario_actual()
    propios = [
        {'ruta': f['ruta'], 'nombre': f['nombre'], 'usuario': int(usuario), 'de': None}
        for f in bd.consultar(
            "SELECT ruta, nombre FROM indice_nombres WHERE usuario_id = %s AND "
            "NOT es_carpeta AND extension IN ('xlsx','xlsm') ORDER BY nombre LIMIT 500",
            (usuario,))
    ]

    ajenos = []
    correo = _correo_usuario(usuario)
    if correo:
        filas = bd.consultar(
            'SELECT DISTINCT propietario_id, ruta FROM compartidos '
            'WHERE LOWER(email) = %s AND propietario_id <> %s '
            'AND (expira_en IS NULL OR expira_en > NOW())', (correo, usuario))
        filas = [f for f in filas if _es_excel(f['ruta'])]
        duenos = tuple({int(f['propietario_id']) for f in filas})
        nombre_de = {}
        if duenos:
            nombre_de = {n['id']: n['nombre'] for n in bd.consultar(
                "SELECT u.id, COALESCE(t.nombres || ' ' || t.apellidos, u.full_name, "
                "u.username) AS nombre FROM usuarios u "
                "LEFT JOIN trabajadores t ON u.trabajador_id = t.id WHERE u.id IN %s",
                (duenos,), nomina=True)}
        for f in filas:
            prop = int(f['propietario_id'])
            try:
                if not os.path.isfile(ruta_fisica(prop, f['ruta'])):
                    continue      # el dueño lo movió o borró
            except Exception:
                continue
            ajenos.append({'ruta': f['ruta'],
                           'nombre': f['ruta'].rsplit('/', 1)[-1],
                           'usuario': prop,
                           'de': nombre_de.get(prop) or ('Usuario %s' % prop)})

    return jsonify({'success': True, 'archivos': propios + ajenos,
                    'propios': len(propios), 'compartidos': len(ajenos)})


@bp_vinculos.route('/vinculos/hojas', methods=['GET'])
def hojas():
    """GET /vinculos/hojas?ruta=&usuario= — hojas de un Excel (propio o compartido)."""
    usuario = usuario_actual()
    try:
        dueno = int(request.args.get('usuario') or usuario)
    except (TypeError, ValueError):
        return error('Dueño inválido', 400)
    try:
        ruta = normalizar_ruta_virtual(request.args.get('ruta', ''))
        if not puede_leer(usuario, dueno, ruta):
            return error('Sin permiso sobre ese archivo', 403)
        fis = ruta_fisica(dueno, ruta)
    except RutaInvalida as excepcion:
        return error(str(excepcion), excepcion.codigo)
    if not _es_excel(ruta):
        return error('No es una hoja de cálculo', 400)
    try:
        wb = openpyxl.load_workbook(fis, read_only=True)
        nombres = list(wb.sheetnames)
        wb.close()
    except Exception as excepcion:
        log.error('hojas %s: %s', ruta, excepcion)
        return error('No se pudo leer el archivo', 500)
    return jsonify({'success': True, 'hojas': nombres})


@bp_vinculos.route('/vinculos/eliminar', methods=['POST'])
def eliminar():
    """POST /vinculos/eliminar {id} — desactiva un vínculo (no borra el valor ya
    escrito en el destino)."""
    usuario = usuario_actual()
    d = request.get_json(silent=True) or {}
    vid = d.get('id')
    if not vid:
        return error('Falta el id del vínculo', 400)
    bd.ejecutar('UPDATE vinculos_datos SET activo = FALSE WHERE id = %s AND '
                'destino_usuario = %s', (vid, usuario))
    return jsonify({'success': True})
