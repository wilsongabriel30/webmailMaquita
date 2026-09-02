# -*- coding: utf-8 -*-
"""
Formularios del Almacén — el Excel de las respuestas, con formato
=================================================================
Hasta el 27/08/2026 la exportación era una rejilla pelada: encabezados sin
distinguir, columnas de ancho por defecto (los textos largos salían cortados) y
las fechas como texto, así que ni siquiera se podían ordenar. Abrirla no decía
gran cosa.

Ahora sale **una hoja y nada más**: una tabla de Excel de verdad —con filtros y
filas alternas—, los encabezados en la primera fila y congelados, y los anchos
calculados. Sin títulos ni recuentos encima, porque estorban a lo que se hace con
este archivo: filtrar, ordenar y copiar rangos.

`hoja_resumen()` se conserva sin usar, por si algún día se quiere volver a
adjuntar el resumen por pregunta; hoy ese resumen se mira en la pantalla de
Respuestas, que además lo enseña con gráficos.

Va en su propio archivo porque el formato de un informe crece cada vez que se
mira: dejarlo dentro de `api_encuestas` habría engordado el módulo de rutas con
algo que no tiene nada que ver con rutas.

Decisiones que conviene recordar:

- **Se usa una Tabla de Excel** (`worksheet.table.Table`), no bordes pintados a
  mano: es lo que da los filtros desplegables y las bandas de color, y OnlyOffice
  —que es con lo que se abre esto en el Drive— la entiende igual.
- **Las fechas van como fecha, no como texto.** Es lo que permite ordenar por
  «cuándo respondió» y filtrar por rango, que es la primera cosa que se hace con
  una tabla de respuestas.
- **Ningún adorno puede impedir la exportación.** Todo lo que es estética va
  dentro de un try/except: si algo falla, el archivo sale igual con los datos,
  que son lo que no se puede perder.

Autoría: Equipo de Tecnología Maquita — 2026-08-27
"""
import logging
import os
import re

log = logging.getLogger('almacen.encuestas.excel')

MORADO = '5B2D8E'
MORADO_SUAVE = 'EDE7F6'
GRIS_TENUE = '5F6368'
BLANCO = 'FFFFFF'

ANCHO_MINIMO = 12
ANCHO_MAXIMO = 55        # más allá, el texto se ajusta en varias líneas
ALTO_TITULO = 30

# Logo institucional. Vive junto al código y no en el Drive: es parte del
# programa, no un archivo de nadie, y así no depende de que exista una ruta ni
# de los permisos de una carpeta.
LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                    'recursos', 'logo-maquita.png')
LOGO_ALTO = 46           # píxeles; el ancho sale de la proporción del original
FILAS_CABECERA = 3       # las que ocupa el logo antes de la tabla


def _fuente(**kw):
    from openpyxl.styles import Font
    return Font(**kw)


def _relleno(color):
    from openpyxl.styles import PatternFill
    return PatternFill('solid', fgColor=color)


def encabezados_unicos(titulos):
    """Excel no admite dos columnas con el mismo nombre dentro de una tabla, y
    dos preguntas SÍ pueden llamarse igual. Se numeran las repetidas en vez de
    fallar o de perder una columna."""
    vistos, salida = {}, []
    for titulo in titulos:
        limpio = (titulo or 'Sin título').strip() or 'Sin título'
        if limpio in vistos:
            vistos[limpio] += 1
            limpio = '%s (%d)' % (limpio, vistos[limpio])
        else:
            vistos[limpio] = 1
        salida.append(limpio)
    return salida


def _ancho_de(texto):
    largo = len(str(texto or ''))
    return max(ANCHO_MINIMO, min(largo + 4, ANCHO_MAXIMO))


def nombre_de_hoja(titulo):
    """Nombre válido para una pestaña de Excel a partir del título.

    Excel no admite `: \\ / ? * [ ]` ni más de 31 caracteres, y una pestaña sin
    nombre no existe. Si no queda nada aprovechable, «Respuestas».
    """
    limpio = re.sub(r'[:\\/?*\[\]]', ' ', str(titulo or '')).strip()
    limpio = re.sub(r'\s+', ' ', limpio)[:31].strip()
    return limpio or 'Respuestas'


def escribir(libro, cabeceras, filas_datos, titulo_formulario, color_tema=None):
    """La hoja de respuestas como tabla. `filas_datos` son listas ya ordenadas."""
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    hoja = libro.active
    hoja.title = nombre_de_hoja(titulo_formulario)
    color = (color_tema or '').lstrip('#').upper() or MORADO
    if not re.fullmatch(r'[0-9A-F]{6}', color):
        color = MORADO

    total_columnas = len(cabeceras)

    # Arriba va el logo institucional y debajo la tabla. La cabecera se reduce a
    # eso: sin título ni recuentos, que es lo que se pidió quitar el 27/08/2026
    # porque estorbaba a filtrar, ordenar y copiar rangos. El nombre del
    # formulario tampoco se pierde: es el de la pestaña y el del archivo.
    #
    # Si el logo no está o no se puede insertar, la tabla sube a la fila 1: el
    # archivo no puede depender de una imagen.
    fila_encabezado = 1
    if _poner_logo(hoja):
        fila_encabezado = FILAS_CABECERA + 1
    for i, titulo in enumerate(cabeceras, start=1):
        celda = hoja.cell(row=fila_encabezado, column=i, value=titulo)
        celda.font = _fuente(bold=True, color=BLANCO)
        celda.fill = _relleno(color)
        celda.alignment = Alignment(vertical='center', wrap_text=True)

    borde = Border(bottom=Side(style='thin', color='D9D9D9'))
    for n, valores in enumerate(filas_datos, start=fila_encabezado + 1):
        for i, valor in enumerate(valores, start=1):
            celda = hoja.cell(row=n, column=i, value=valor)
            celda.alignment = Alignment(vertical='top', wrap_text=True)
            celda.border = borde
            # La fecha va en la primera columna y se escribe como fecha real;
            # aquí solo se le pone el formato con el que se lee.
            if i == 1 and hasattr(valor, 'strftime'):
                celda.number_format = 'DD/MM/YYYY HH:MM'

    try:
        _tabla(hoja, fila_encabezado, total_columnas, len(filas_datos))
    except Exception as excepcion:      # el archivo sale igual, sin la tabla
        log.warning('excel: no se pudo dar formato de tabla (%s)', excepcion)
        hoja.auto_filter.ref = '%s%d:%s%d' % (
            'A', fila_encabezado, get_column_letter(max(total_columnas, 1)),
            fila_encabezado + len(filas_datos))

    # Anchos: se mira el encabezado y las primeras filas, no todas: con mil
    # respuestas recorrerlas enteras para calcular un ancho no compensa.
    for i, titulo in enumerate(cabeceras, start=1):
        ancho = _ancho_de(titulo)
        for valores in filas_datos[:60]:
            if i <= len(valores):
                ancho = max(ancho, _ancho_de(valores[i - 1]))
        hoja.column_dimensions[get_column_letter(i)].width = ancho

    # Encabezados siempre a la vista al bajar por las respuestas.
    hoja.freeze_panes = hoja.cell(row=fila_encabezado + 1, column=1)
    return hoja


def _poner_logo(hoja):
    """Coloca el logo en la esquina superior izquierda. Devuelve si se puso.

    Se ancla en A1 y las filas de arriba se dejan con la altura justa, para que
    no quede flotando sobre los datos ni empuje la tabla más de lo necesario.
    """
    try:
        if not os.path.isfile(LOGO):
            log.warning('excel: no está el logo en %s', LOGO)
            return False
        from openpyxl.drawing.image import Image as ImagenExcel

        imagen = ImagenExcel(LOGO)
        # Se respeta la proporción del original: un logo estirado se ve peor que
        # no ponerlo.
        proporcion = (imagen.width / imagen.height) if imagen.height else 3.3
        imagen.height = LOGO_ALTO
        imagen.width = int(LOGO_ALTO * proporcion)
        imagen.anchor = 'A1'
        hoja.add_image(imagen)

        # Alto de las filas de la cabecera (en puntos: 1 px ≈ 0,75 pt).
        hoja.row_dimensions[1].height = LOGO_ALTO * 0.75
        for n in range(2, FILAS_CABECERA + 1):
            hoja.row_dimensions[n].height = 6
        return True
    except Exception as excepcion:
        # Sin Pillow, con un PNG corrupto o con una versión de openpyxl que no
        # admita imágenes, el Excel sale igual: solo se queda sin logo.
        log.warning('excel: no se pudo poner el logo (%s)', excepcion)
        return False


def _tabla(hoja, fila_encabezado, columnas, cuantas_filas):
    """Convierte el rango en una Tabla de Excel (filtros + filas alternas)."""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    if not columnas:
        return
    # Una tabla necesita al menos una fila bajo el encabezado; sin respuestas se
    # deja solo el filtro, que no exige datos.
    ultima = fila_encabezado + max(cuantas_filas, 1)
    referencia = 'A%d:%s%d' % (fila_encabezado,
                               get_column_letter(columnas), ultima)
    tabla = Table(displayName='Respuestas', ref=referencia)
    tabla.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium9', showRowStripes=True, showColumnStripes=False,
        showFirstColumn=False, showLastColumn=False)
    hoja.add_table(tabla)


def hoja_resumen(libro, resumen_por_pregunta, color_tema=None):
    """Segunda hoja: cuántas veces se eligió cada opción y el promedio de las
    escalas. Es lo que se mira antes de leer respuesta por respuesta."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    color = (color_tema or '').lstrip('#').upper() or MORADO
    if not re.fullmatch(r'[0-9A-F]{6}', color):
        color = MORADO

    hoja = libro.create_sheet('Resumen')
    hoja.cell(row=1, column=1, value='Resumen de las respuestas')
    hoja.cell(row=1, column=1).font = _fuente(size=14, bold=True, color=BLANCO)
    hoja.cell(row=1, column=1).fill = _relleno(color)
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    hoja.row_dimensions[1].height = ALTO_TITULO

    n = 3
    for bloque in resumen_por_pregunta:
        hoja.cell(row=n, column=1, value=bloque['titulo'])
        hoja.cell(row=n, column=1).font = _fuente(bold=True, size=11)
        hoja.cell(row=n, column=1).fill = _relleno(MORADO_SUAVE)
        hoja.merge_cells(start_row=n, start_column=1, end_row=n, end_column=3)
        n += 1

        if bloque.get('promedio') is not None:
            hoja.cell(row=n, column=1, value='Promedio')
            hoja.cell(row=n, column=2, value=round(bloque['promedio'], 2))
            n += 1

        for etiqueta, veces, porcentaje in bloque.get('conteo') or []:
            hoja.cell(row=n, column=1, value=etiqueta)
            hoja.cell(row=n, column=2, value=veces)
            celda = hoja.cell(row=n, column=3, value=porcentaje / 100.0)
            celda.number_format = '0.0%'
            n += 1

        if bloque.get('respondidas') is not None and not bloque.get('conteo'):
            hoja.cell(row=n, column=1, value='Respuestas escritas')
            hoja.cell(row=n, column=2, value=bloque['respondidas'])
            n += 1
        n += 1      # una fila en blanco entre preguntas

    hoja.column_dimensions['A'].width = 52
    hoja.column_dimensions['B'].width = 12
    hoja.column_dimensions['C'].width = 12
    for fila in hoja.iter_rows(min_row=3, max_row=max(n, 3), min_col=1, max_col=1):
        for celda in fila:
            celda.alignment = Alignment(wrap_text=True, vertical='center')
    return hoja
